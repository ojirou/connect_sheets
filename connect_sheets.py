import openpyxl
wb = openpyxl.load_workbook(".//data_connect_sheets//sample.xlsx")
# 1つ目のシートかどうか
is_first_sheet = True
# このリストに全シートから読み取ったデータをまとめる
row_list = []
# ブックの中のすべてのシートを処理
for ws in wb.worksheets:
    if is_first_sheet:
        start_row = 1
    else:
        start_row = 5
    for row in ws.iter_rows(min_row=start_row):
        # ヘッダーより下で空行になったら読み込み終了
        if row[8].row > 4 and row[8].value is None:
            break
        if row[1].value is None:
            continue
        value_list = []
        for c in row:
            value_list.append(c.value)
#       if row[2] is not None:
        row_list.append(value_list)
        is_first_sheet = False
# データを転記する新しいシート
ws_new = wb.create_sheet(title="Summary", index=0)
# 書き込み時の行番号
row_num = 1
# 新しいシートに1行ずつデータを書き込む
for row in row_list:
    # 1行分のデータを書き込む
#    if row[2] is not None:
    ws_new.append(row)
    ws_new.cell(row_num,5).value = ""
    ws_new.cell(row_num,6).value = ""
    ws_new.cell(row_num,7).value = ""
    ws_new.cell(row_num,8).value = ""
    ws_new.cell(row_num,9).value = ""
    ws_new.cell(row_num,10).value = ""
    # データ部分のA列に日付の表示形式を設定し、F列の数式を書き換え
    if row_num > 1:
        ws_new.cell(row_num, 2).number_format = "yyyy/m/d"
        price_yen=row[4]
#        price_yen=str(price_yen)
#        price_yen.replace("円","")
        ws_new.cell(row_num, 4).value = price_yen.replace("円","")
#        ws_new.cell(row_num, 6).value = "=D" + str(row_num) + "*E" + str(row_num)
    row_num = row_num + 1
#wb.move_sheet(ws_new, )
wb.save(".//data_connect_sheets//sample.xlsx")